#!/usr/bin/env python3
"""
Process Unified Response.xlsx — extract emails with "Read" status
and update campaign_leads in D1 database.

What it does:
  1. Finds all emails with "Read" in the campaign column
  2. Updates campaign_leads: response='Read', email_variant_sent='office', sent=1
  3. Trims the email field to just the office@ address (removes other variants)

Usage:
  python3 tools/process-email-responses.py                  # Dry run (default)
  python3 tools/process-email-responses.py --apply-local    # Apply to local dev DB
  python3 tools/process-email-responses.py --apply-remote   # Apply to live DB

Source: ~/Desktop/kiro/emails/Response/Unified Response.xlsx
Target column: 📮 School Campaign (2026/07/03)
"""

import openpyxl
import subprocess
import sys
import os

# Config
XLSX_PATH = os.path.expanduser('~/Desktop/kiro/emails/Response/Unified Response.xlsx')
CAMPAIGN_COL_NAME = '📮 School Campaign (2026/07/03)'
SWF_DIR = os.path.expanduser('~/Code/xConvert.com/scrabblewordsfinder')


def load_read_emails():
    """Parse all sheets, find emails with 'Read' status in the campaign column."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
    read_emails = set()
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = rows[0]

        # Find the email column and campaign column
        email_col = None
        campaign_col = None
        for i, h in enumerate(headers):
            if h and 'email' in str(h).lower():
                email_col = i
            if h and CAMPAIGN_COL_NAME in str(h):
                campaign_col = i

        if email_col is None or campaign_col is None:
            print(f"  ⚠️  {sheet_name}: Missing columns (email={email_col}, campaign={campaign_col})")
            continue

        sheet_reads = 0
        for row in rows[1:]:
            total_rows += 1
            if len(row) > campaign_col and row[campaign_col]:
                status = str(row[campaign_col]).strip()
                email = str(row[email_col]).strip().lower() if row[email_col] else ''
                if status == 'Read' and email:
                    # Extract just the office@ address if multiple are present
                    if ',' in email:
                        parts = [e.strip() for e in email.split(',')]
                        office = next((e for e in parts if e.startswith('office@')), parts[0])
                        read_emails.add(office)
                    else:
                        read_emails.add(email)
                    sheet_reads += 1

        print(f"  📄 {sheet_name}: {sheet_reads} reads found")

    wb.close()
    print(f"\n  Total rows processed: {total_rows}")
    print(f"  Unique emails with 'Read' status: {len(read_emails)}")
    return read_emails


def generate_sql(read_emails):
    """Generate UPDATE SQL statements for emails that were read.
    
    For each read email:
    - Set response = 'Read'
    - Set email_variant_sent = 'office' (we only sent to office@ addresses)
    - Set sent = 1 (the email was definitely sent)
    - Trim email field to just the office@ address
    """
    statements = []
    for email in sorted(read_emails):
        safe_email = email.replace("'", "''")
        # Update response, variant, sent status, AND trim email to office@ only
        statements.append(
            f"UPDATE campaign_leads SET response = 'Read', email_variant_sent = 'office', sent = 1, email = '{safe_email}' WHERE LOWER(email) LIKE '%{safe_email}%';"
        )
    return statements


def apply_to_db(statements, remote=False):
    """Apply SQL statements to D1 database."""
    sql_file = '/tmp/campaign_read_updates.sql'
    with open(sql_file, 'w') as f:
        f.write('\n'.join(statements))

    cmd = ['npx', 'wrangler', 'd1', 'execute', 'DB']
    if remote:
        cmd.append('--remote')
    else:
        cmd.append('--local')
    cmd.extend(['--file', sql_file])

    print(f"\n  Running: {' '.join(cmd)}")
    print(f"  SQL file: {sql_file} ({len(statements)} statements)")

    result = subprocess.run(
        cmd, cwd=SWF_DIR, capture_output=True, text=True,
        input='Y\n'  # Auto-confirm large file prompt
    )

    if result.returncode == 0:
        print("  ✅ Applied successfully!")
        if 'rows_written' in result.stdout:
            for line in result.stdout.split('\n'):
                if 'rows_written' in line:
                    print(f"  {line.strip()}")
    else:
        print(f"  ❌ Error: {result.stderr[-500:]}")


def main():
    mode = '--dry-run'
    if len(sys.argv) > 1:
        mode = sys.argv[1]

    print("=" * 60)
    print("📮 Process Email Responses — Campaign Leads Read Status")
    print("=" * 60)
    print(f"\n  Source: {XLSX_PATH}")
    print(f"  Column: {CAMPAIGN_COL_NAME}")
    print(f"  Mode:   {mode}")
    print()

    # Step 1: Parse XLSX
    print("Step 1: Parsing XLSX...")
    read_emails = load_read_emails()

    if not read_emails:
        print("\n  ⚠️  No 'Read' emails found. Nothing to update.")
        return

    # Step 2: Generate SQL
    print("\nStep 2: Generating SQL...")
    statements = generate_sql(read_emails)
    print(f"  Generated {len(statements)} UPDATE statements")
    print(f"  Each sets: response='Read', email_variant_sent='office', sent=1, email=office@ only")

    # Step 3: Show sample
    print("\nSample (first 3):")
    for s in statements[:3]:
        print(f"  {s}")
    if len(statements) > 3:
        print(f"  ... and {len(statements) - 3} more")

    # Step 4: Apply based on mode
    if mode == '--dry-run':
        print("\n" + "=" * 60)
        print("🔍 DRY RUN — No changes applied.")
        print("=" * 60)
        print("\nTo apply:")
        print(f"  python3 tools/process-email-responses.py --apply-local   # Local dev DB")
        print(f"  python3 tools/process-email-responses.py --apply-remote  # Live DB")
    elif mode == '--apply-local':
        print("\n  Applying to LOCAL dev database...")
        apply_to_db(statements, remote=False)
    elif mode == '--apply-remote':
        print("\n  Applying to REMOTE (live) database...")
        apply_to_db(statements, remote=True)
    else:
        print(f"\n  ❌ Unknown mode: {mode}")
        print("  Use: --dry-run, --apply-local, or --apply-remote")


if __name__ == '__main__':
    main()
